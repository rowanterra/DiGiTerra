"""
Shared helpers for preprocessing, prediction, and Excel export. VIS_DIR (from config)
is where we write predictions.csv and similar. Handoff: see python_scripts/config.py
and HANDOFF.md for paths.
"""
import os
import logging
from pathlib import Path

import xlsxwriter
import openpyxl
import pandas as pd
import numpy as np
from sklearn.base import clone
from sklearn.compose import TransformedTargetRegressor
from sklearn.model_selection import (
    KFold,
    StratifiedKFold,
    RepeatedKFold,
    RepeatedStratifiedKFold,
    ShuffleSplit,
    StratifiedShuffleSplit,
    cross_validate,
)
from sklearn.pipeline import Pipeline

from python_scripts.preprocessing.utilites import make_preprocessor, _get_scaler

from python_scripts.config import VIS_DIR

logger = logging.getLogger(__name__)


def unpack_classification_result(result_tuple):
    """
    Helper function to unpack classification model results with backward compatibility.
    Handles both old format (8 values) and new format (9 values with additional_metrics).
    
    Returns:
        tuple: (report, cm, params, shapes, storedModel, X_scaler, quantileBin_results, feature_order, additional_metrics)
    """
    if len(result_tuple) >= 9:
        return result_tuple[:9]
    else:
        # Old format - pad with None for additional_metrics
        return result_tuple + (None,)


def preprocess_data(
    df,
    target_cols=None,
    indicator_cols=None,
    drop_missing='all',  # 'all', 'indicator', 'target' should be a check box or something maybe?
    impute_strategy=None,  # None, 'mean', 'median', 'knn', or 0 or 0.01
    drop_zero='all',
    zero_replace_values=None,  # list of exact values to convert to 0
    zero_replace_patterns=None  # list of lambda or callable to match patterns (e.g., lambda x: str(x).startswith('<'))
):
    df_clean = df.copy()


#if no impute strategy -> drop rows indicated by user    
    if impute_strategy=='none':
        if drop_missing == 'all':
            df_clean.dropna(inplace=True)
        elif drop_missing == 'indicator':
            #if one indicator missing drop the row
            df_clean.dropna(subset=indicator_cols, how='any', inplace=True)
        elif drop_missing == 'target' and target_cols:
            #if one target missing drop the row
            df_clean.dropna(subset=target_cols, inplace=True, how='any')
        elif drop_missing == 'indicatorAndTarget' and target_cols:
            #if one target or indicator missing drop the row
            df_clean.dropna(subset=indicator_cols + target_cols, inplace=True, how='any')

#if impute strategy -> replace missing values 
    else:
        if drop_missing == 'all':
            for col in df_clean.select_dtypes(include=[np.number]).columns:
                if impute_strategy == 'mean':
                    df_clean[col].fillna(df_clean[col].mean(), inplace=True)
                elif impute_strategy == 'median':
                    df_clean[col].fillna(df_clean[col].median(), inplace=True)
                elif isinstance(impute_strategy, (int, float)):
                    df_clean[col].fillna(impute_strategy, inplace=True)

        elif drop_missing == 'indicator':
            for col in df_clean[indicator_cols].select_dtypes(include=[np.number]).columns:
                if impute_strategy == 'mean':
                    df_clean[col].fillna(df_clean[col].mean(), inplace=True)
                elif impute_strategy == 'median':
                    df_clean[col].fillna(df_clean[col].median(), inplace=True)
                elif isinstance(impute_strategy, (int, float)):
                    df_clean[col].fillna(impute_strategy, inplace=True)

        elif drop_missing == 'target' and target_cols:
            for col in df_clean[target_cols].select_dtypes(include=[np.number]).columns:
                if impute_strategy == 'mean':
                    df_clean[col].fillna(df_clean[col].mean(), inplace=True)
                elif impute_strategy == 'median':
                    df_clean[col].fillna(df_clean[col].median(), inplace=True)
                elif isinstance(impute_strategy, (int, float)):
                    df_clean[col].fillna(impute_strategy, inplace=True)

        elif drop_missing == 'indicatorAndTarget' and target_cols:
            for col in df_clean[indicator_cols+target_cols].select_dtypes(include=[np.number]).columns:
                if impute_strategy == 'mean':
                    df_clean[col].fillna(df_clean[col].mean(), inplace=True)
                elif impute_strategy == 'median':
                    df_clean[col].fillna(df_clean[col].median(), inplace=True)
                elif isinstance(impute_strategy, (int, float)):
                    df_clean[col].fillna(impute_strategy, inplace=True)

    

#drop rows that are zeros in the columns indicated 
    if drop_zero == 'all':
        df_clean = df_clean[(df_clean != 0).any(axis=1)]
    elif drop_zero == 'indicator':
        df_clean = df_clean[~(df_clean[indicator_cols].eq(0).all(axis=1))]
    elif drop_zero == 'target' and target_cols:
        df_clean = df_clean[~(df_clean[target_cols].eq(0).all(axis=1))]
    elif drop_zero == 'indicatorAndTarget' and target_cols:
        subset = indicator_cols + target_cols
        df_clean = df_clean[~(df_clean[subset].eq(0).all(axis=1))]

    # ---- Step 2: Replace values with 0 ----
    if zero_replace_values:
        df_clean.replace(zero_replace_values, 0, inplace=True)

    if zero_replace_patterns:
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].apply(
                lambda x: 0 if any([pattern(x) for pattern in zero_replace_patterns]) else x
            )

    return df_clean


def write_to_excel(data, indicator_names, predictor_names, stratify_name, modelName, params, units, trainOverall, testOverall, train_results, test_results, scaler, seed, shapes, quantileBin_results, cross_validation_summary=None, feature_selection_info=None, outlier_info=None):
        workbook = xlsxwriter.Workbook(str(VIS_DIR / "model_performance.xlsx"), {'nan_inf_to_errors': True})
        worksheet = workbook.add_worksheet('Preprocessing')
        worksheet.write('A1', 'Filename')
        worksheet.write('B1', data['filename'])
        worksheet.write('A2', 'Indicators')
        indicator_list = indicator_names.tolist() if hasattr(indicator_names, 'tolist') else list(indicator_names)
        worksheet.write('B2', f'{[str(i)[:10] for i in indicator_list]}')
        worksheet.write('A3', 'Predictors')
        predictor_list = predictor_names.tolist() if hasattr(predictor_names, 'tolist') else list(predictor_names)
        worksheet.write('B3', f'{[str(p)[:10] for p in predictor_list]}')
        worksheet.write('A4', 'Stratify Variable')
        worksheet.write('B4', stratify_name)
        worksheet.write('A5', 'Scaling')
        worksheet.write('B5', scaler)
        worksheet.write('A6', 'Seed')
        worksheet.write('B6', seed)

        worksheet.write('A7', 'X_train shape:')
        worksheet.write('B7', str(shapes['X_train']))
        worksheet.write('A8', 'X_test shape:')
        worksheet.write('B8', str(shapes['X_test']))
        worksheet.write('A9', 'y_train shape:')
        worksheet.write('B9', str(shapes['y_train']))
        worksheet.write('A10', 'y_test shape:')
        worksheet.write('B10', str(shapes['y_test']))
        worksheet.write('A11', 'train_groups')
        # train_groups.to_frame()
        # with pd.ExcelWriter('static/performance.xlsx', engine='xlsxwriter') as writer:
        #     train_groups.to_excel(writer, sheet_name='quantileBins', startrow=11, index=True, header=True)
        

        worksheet2 = workbook.add_worksheet('Model Selection')
        worksheet2.write('A1', 'Model')
        worksheet2.write('B1', f'{modelName}')
        worksheet2.write('A2', 'Hyperparameters')
        
        row = 3  # B2 means row 1 (0-indexed) and column 1 (also 0-indexed)
        col = 1  # Column B

        worksheet2.write(row - 1, col, "Parameter")  # Header
        worksheet2.write(row - 1, col + 1, "Value")  # Header

        for key, value in params.items():
            worksheet2.write(row, col, key)
            worksheet2.write(row, col + 1, str(value))  # Convert to string just in case
            row += 1

        worksheet3 = workbook.add_worksheet('Overall Performance')

        header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC'})
        worksheet3.write(0, 0, 'Units:', header_format)
        worksheet3.write(0, 1, units)

        worksheet3.write('A2', 'Training Score')
        worksheet3.write('B2', trainOverall['R²'])
        worksheet3.write('A3', 'Validation Score')
        worksheet3.write('B3', testOverall['R²'])

        worksheet3.write('A4', 'Training RMSE')
        worksheet3.write('B4', trainOverall['RMSE'])
        worksheet3.write('A5', 'Validation RMSE')
        worksheet3.write('B5', testOverall['RMSE'])

        worksheet3.write('A6', 'Training MAE')
        worksheet3.write('B6', trainOverall['MAE'])
        worksheet3.write('A7', 'Validation MAE')
        worksheet3.write('B7', testOverall['MAE'])
        
        # Add additional regression metrics if available
        row = 8
        if 'Explained Variance' in trainOverall.index:
            worksheet3.write(f'A{row}', 'Training Explained Variance')
            worksheet3.write(f'B{row}', trainOverall['Explained Variance'])
            row += 1
            worksheet3.write(f'A{row}', 'Validation Explained Variance')
            worksheet3.write(f'B{row}', testOverall['Explained Variance'])
            row += 1
        
        if 'Median AE' in trainOverall.index:
            worksheet3.write(f'A{row}', 'Training Median AE')
            worksheet3.write(f'B{row}', trainOverall['Median AE'])
            row += 1
            worksheet3.write(f'A{row}', 'Validation Median AE')
            worksheet3.write(f'B{row}', testOverall['Median AE'])
            row += 1
        
        if 'Max Error' in trainOverall.index:
            worksheet3.write(f'A{row}', 'Training Max Error')
            worksheet3.write(f'B{row}', trainOverall['Max Error'])
            row += 1
            worksheet3.write(f'A{row}', 'Validation Max Error')
            worksheet3.write(f'B{row}', testOverall['Max Error'])
            row += 1
        
        if 'MAPE' in trainOverall.index:
            worksheet3.write(f'A{row}', 'Training MAPE')
            worksheet3.write(f'B{row}', trainOverall['MAPE'])
            row += 1
            worksheet3.write(f'A{row}', 'Validation MAPE')
            worksheet3.write(f'B{row}', testOverall['MAPE'])
        
        workbook.close()
        file_path = VIS_DIR / "model_performance.xlsx"

        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            train_results.to_excel(writer, sheet_name='Train Results', index=False)
            test_results.to_excel(writer, sheet_name='Test Results', index=False)

            if (isinstance(quantileBin_results, pd.DataFrame)):
                logger.debug(f"Quantile bin results:\n{quantileBin_results}")
                quantileBin_results.to_excel(writer, sheet_name='quantile Bin Results', index=False)
            
            # Add Cross Validation sheet if available
            if cross_validation_summary and len(cross_validation_summary) > 0:
                cv_df = pd.DataFrame(cross_validation_summary)
                cv_df.to_excel(writer, sheet_name='Cross Validation', index=False)
            
            # Add Feature Selection sheet if available
            if feature_selection_info:
                fs_data = {
                    'Property': ['Method', 'K Requested', 'Original Features', 'Selected Features', 'Selected Feature Names'],
                    'Value': [
                        feature_selection_info.get('method', 'N/A'),
                        feature_selection_info.get('k_requested', 'N/A'),
                        feature_selection_info.get('original_count', 'N/A'),
                        feature_selection_info.get('selected_count', 'N/A'),
                        ', '.join(feature_selection_info.get('selected_features', [])) if feature_selection_info.get('selected_features') else 'N/A'
                    ]
                }
                fs_df = pd.DataFrame(fs_data)
                fs_df.to_excel(writer, sheet_name='Feature Selection', index=False)
            
            # Add Outlier Handling sheet if available
            if outlier_info:
                oi_data = {
                    'Property': ['Method', 'Action', 'Outliers Detected', 'Original Samples', 'Remaining Samples'],
                    'Value': [
                        outlier_info.get('method', 'N/A'),
                        outlier_info.get('action', 'N/A'),
                        outlier_info.get('n_outliers', 0),
                        outlier_info.get('original_samples', 'N/A'),
                        outlier_info.get('remaining_samples', 'N/A')
                    ]
                }
                oi_df = pd.DataFrame(oi_data)
                oi_df.to_excel(writer, sheet_name='Outlier Handling', index=False)


def write_to_excelClassifier(data, indicator_names, predictor_names, stratify_name, scaler, seed, modelName, params, units, report, cm, additional_metrics=None):
    """
    Write classification results to Excel with comprehensive metrics.
    
    Args:
        additional_metrics: Dictionary with additional metrics like ROC AUC, Average Precision, etc.
                          from calculate_classification_metrics()
    """
    workbook = xlsxwriter.Workbook(str(VIS_DIR / "model_performance.xlsx"), {'nan_inf_to_errors': True})
    worksheet = workbook.add_worksheet('Preprocessing')
    worksheet.write('A1', 'Filename')
    worksheet.write('B1', data['filename'])
    worksheet.write('A2', 'Indicators')
    indicator_list = indicator_names.tolist() if hasattr(indicator_names, 'tolist') else list(indicator_names)
    worksheet.write('B2', f'{[str(i)[:10] for i in indicator_list]}')
    worksheet.write('A3', 'Predictors')
    predictor_list = predictor_names.tolist() if hasattr(predictor_names, 'tolist') else list(predictor_names)
    worksheet.write('B3', f'{[str(p)[:10] for p in predictor_list]}')
    worksheet.write('A4', 'Stratify Variable')
    worksheet.write('B4', stratify_name)
    worksheet.write('A5', 'Scaling')
    worksheet.write('B5', scaler)
    worksheet.write('A6', 'Seed')
    worksheet.write('B6', seed)

    worksheet2 = workbook.add_worksheet('Model Selection')
    worksheet2.write('A1', 'Model')
    worksheet2.write('B1', f'{modelName}')
    worksheet2.write('A2', 'Hyperparameters')
    
    row = 3
    col = 1

    worksheet2.write(row - 1, col, "Parameter")
    worksheet2.write(row - 1, col + 1, "Value")

    for key, value in params.items():
        worksheet2.write(row, col, key)
        worksheet2.write(row, col + 1, str(value))
        row += 1

    worksheet3 = workbook.add_worksheet('Overall Performance')
    header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC'})
    worksheet3.write(0, 0, 'Units:', header_format)
    worksheet3.write(0, 1, units)

    row = 1
    col = 0
    logger.debug("Writing classification results to Excel")
    for key, value in report.items():
        worksheet3.write(row, col, key)
        if key=='accuracy':
            worksheet3.write(row, col+1, value)
        else:
            for key2,value2 in value.items():
                worksheet3.write(row, col+1, key2)
                worksheet3.write(row, col+2, value2)
                row+=1
        row += 1

    # Add confusion matrix
    col=1
    worksheet3.write(row,0, 'Confusion Matrix')
    for i in range(cm.shape[0]):      
        for j in range(cm.shape[1]):  
            worksheet3.write(row+i+1, col+j, cm[i, j])
    
    row += cm.shape[0] + 2

    # Add additional metrics if provided
    if additional_metrics:
        worksheet3.write(row, 0, 'Additional Metrics')
        row += 1
        
        # Basic metrics
        if 'accuracy' in additional_metrics:
            worksheet3.write(row, 0, 'Accuracy')
            worksheet3.write(row, 1, additional_metrics['accuracy'])
            row += 1
        
        # Precision, Recall, F1 (macro, micro, weighted)
        for metric_type in ['precision', 'recall', 'f1']:
            for avg_type in ['macro', 'micro', 'weighted']:
                key = f'{metric_type}_{avg_type}'
                if key in additional_metrics:
                    worksheet3.write(row, 0, f'{metric_type.capitalize()} ({avg_type})')
                    worksheet3.write(row, 1, additional_metrics[key])
                    row += 1
        
        # ROC AUC
        if 'roc_auc' in additional_metrics:
            worksheet3.write(row, 0, 'ROC AUC')
            worksheet3.write(row, 1, additional_metrics['roc_auc'])
            row += 1
        for avg_type in ['macro', 'micro', 'weighted']:
            key = f'roc_auc_{avg_type}'
            if key in additional_metrics:
                worksheet3.write(row, 0, f'ROC AUC ({avg_type})')
                worksheet3.write(row, 1, additional_metrics[key])
                row += 1
        
        # Average Precision
        if 'average_precision' in additional_metrics:
            worksheet3.write(row, 0, 'Average Precision')
            worksheet3.write(row, 1, additional_metrics['average_precision'])
            row += 1
        for avg_type in ['macro', 'micro', 'weighted']:
            key = f'average_precision_{avg_type}'
            if key in additional_metrics:
                worksheet3.write(row, 0, f'Average Precision ({avg_type})')
                worksheet3.write(row, 1, additional_metrics[key])
                row += 1

    workbook.close()
    file_path = VIS_DIR / "model_performance.xlsx"
    
    # Use openpyxl to add additional sheets with detailed metrics
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        # Classification Report as DataFrame
        if isinstance(report, dict):
            report_data = []
            for key, value in report.items():
                if key == 'accuracy':
                    report_data.append({'Metric': 'accuracy', 'Value': value})
                else:
                    for key2, value2 in value.items():
                        report_data.append({'Metric': f'{key}_{key2}', 'Value': value2})
            if report_data:
                report_df = pd.DataFrame(report_data)
                report_df.to_excel(writer, sheet_name='Classification Report', index=False)
        
        # Confusion Matrix as DataFrame
        cm_df = pd.DataFrame(cm)
        cm_df.to_excel(writer, sheet_name='Confusion Matrix', index=False)
        
        # Additional Metrics as DataFrame
        if additional_metrics:
            metrics_data = []
            # Basic metrics
            for key in ['accuracy', 'precision_macro', 'precision_micro', 'precision_weighted',
                       'recall_macro', 'recall_micro', 'recall_weighted',
                       'f1_macro', 'f1_micro', 'f1_weighted']:
                if key in additional_metrics:
                    metrics_data.append({'Metric': key.replace('_', ' ').title(), 'Value': additional_metrics[key]})
            
            # ROC AUC metrics
            for key in ['roc_auc', 'roc_auc_macro', 'roc_auc_micro', 'roc_auc_weighted']:
                if key in additional_metrics:
                    metrics_data.append({'Metric': key.replace('_', ' ').upper(), 'Value': additional_metrics[key]})
            
            # Average Precision metrics
            for key in ['average_precision', 'average_precision_macro', 'average_precision_micro', 'average_precision_weighted']:
                if key in additional_metrics:
                    metrics_data.append({'Metric': key.replace('_', ' ').title(), 'Value': additional_metrics[key]})
            
            if metrics_data:
                metrics_df = pd.DataFrame(metrics_data)
                metrics_df.to_excel(writer, sheet_name='Additional Metrics', index=False)
            
            # Per-class metrics
            if 'per_class' in additional_metrics:
                per_class_data = []
                for key, value in additional_metrics['per_class'].items():
                    per_class_data.append({'Metric': key.replace('_', ' ').title(), 'Value': value})
                if per_class_data:
                    per_class_df = pd.DataFrame(per_class_data)
                    per_class_df.to_excel(writer, sheet_name='Per-Class Metrics', index=False)
            
            if 'per_class_roc_auc' in additional_metrics:
                roc_auc_data = []
                for key, value in additional_metrics['per_class_roc_auc'].items():
                    roc_auc_data.append({'Metric': key.replace('_', ' ').upper(), 'Value': value})
                if roc_auc_data:
                    roc_auc_df = pd.DataFrame(roc_auc_data)
                    roc_auc_df.to_excel(writer, sheet_name='Per-Class ROC AUC', index=False)
            
            if 'per_class_ap' in additional_metrics:
                ap_data = []
                for key, value in additional_metrics['per_class_ap'].items():
                    ap_data.append({'Metric': key.replace('_', ' ').title(), 'Value': value})
                if ap_data:
                    ap_df = pd.DataFrame(ap_data)
                    ap_df.to_excel(writer, sheet_name='Per-Class Average Precision', index=False)

    
def write_to_excelCluster(data, indicator_names, stratify_name, scaler, seed, modelName, params, units, train_silhouette, train_calinski_harabasz, train_davies_bouldin, test_silhouette, test_calinski_harabasz, test_davies_bouldin, best_k, centers, silhouette_grid):
    
    logger.debug(f"Cluster centers:\n{centers}")
    logger.debug(f"Silhouette grid:\n{silhouette_grid}")
    workbook = xlsxwriter.Workbook(str(VIS_DIR / "model_performance.xlsx"), {'nan_inf_to_errors': True})
    worksheet = workbook.add_worksheet('Preprocessing')
    worksheet.write('A1', 'Filename')
    worksheet.write('B1', data['filename'])
    worksheet.write('A2', 'Indicators')
    indicator_list = indicator_names.tolist() if hasattr(indicator_names, 'tolist') else list(indicator_names)
    worksheet.write('B2', f'{[str(i)[:10] for i in indicator_list]}')
    worksheet.write('A3', 'Stratify Variable')
    worksheet.write('B3', stratify_name)
    worksheet.write('A4', 'Scaling')
    worksheet.write('B4', scaler)
    worksheet.write('A5', 'Seed')
    worksheet.write('B5', seed)


    worksheet2 = workbook.add_worksheet('Model Selection')
    worksheet2.write('A1', 'Model')
    worksheet2.write('B1', f'{modelName}')
    worksheet2.write('A2', 'Hyperparameters')
    
    row = 3  # B2 means row 1 (0-indexed) and column 1 (also 0-indexed)
    col = 1  # Column B

    worksheet2.write(row - 1, col, "Parameter")  # Header
    worksheet2.write(row - 1, col + 1, "Value")  # Header

    for key, value in params.items():
        worksheet2.write(row, col, key)
        worksheet2.write(row, col + 1, str(value))  # Convert to string just in case
        row += 1

    worksheet3 = workbook.add_worksheet('Overall Performance')
    header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC'})
    worksheet3.write(0, 0, 'Units:', header_format)
    worksheet3.write(0, 1, units)

    worksheet3.write('A2', 'Train Silhouette')
    worksheet3.write('B2', train_silhouette)
    worksheet3.write('A3', 'Train Calinski Harabasz')
    worksheet3.write('B3', train_calinski_harabasz)
    worksheet3.write('A4', 'Train Davies Bouldin')
    worksheet3.write('B4', train_davies_bouldin)

    worksheet3.write('A5', 'Test Silhouette')
    worksheet3.write('B5', test_silhouette)
    worksheet3.write('A6', 'Test Calinski Harabasz')
    worksheet3.write('B6', test_calinski_harabasz)
    worksheet3.write('A7', 'Test Davies Bouldin')
    worksheet3.write('B7', test_davies_bouldin)

    worksheet3.write('A8', 'Best K')
    worksheet3.write('B8', best_k)

    centersrow=8
    centerscol=1
    worksheet3.write('A9', 'Centers')
    if centers is not None and hasattr(centers, "shape"):
        for i in range(centers.shape[0]):      
            for j in range(centers.shape[1]):  
                worksheet3.write(centersrow+i, centerscol+j, centers[i, j])
        gridrow=centersrow+i+1
    else:
        worksheet3.write(centersrow, centerscol, "N/A")
        gridrow=centersrow+1

    gridcol=1
    worksheet3.write(gridrow,0,'Silhouette Grid')
    gridrow+=1
    for key, value in silhouette_grid.items():
        worksheet3.write(gridrow, gridcol, key)
        worksheet3.write(gridrow, gridcol + 1, str(value))  # Convert to string just in case
        gridrow += 1

    workbook.close()
    file_path = VIS_DIR / "model_performance.xlsx"


def _build_cv_summary(results: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    fold_df = pd.DataFrame(results)
    summary_rows = []
    for key, values in results.items():
        if not key.startswith("test_"):
            continue
        metric_name = key.replace("test_", "")
        metric_values = np.array(values, dtype=float)
        if metric_name.startswith("neg_"):
            metric_name = metric_name.replace("neg_", "")
            metric_values = -metric_values
        summary_rows.append({
            "Metric": metric_name,
            "Mean": float(np.mean(metric_values)),
            "Std": float(np.std(metric_values)),
        })
    summary_df = pd.DataFrame(summary_rows)
    return summary_df, fold_df


def write_cross_validation_results(summary_df: pd.DataFrame, fold_df: pd.DataFrame, cv_type: str, cv_folds: int) -> Path:
    file_path = VIS_DIR / "cross_validation.xlsx"
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        fold_df.to_excel(writer, sheet_name="Fold Scores", index=False)
        meta_df = pd.DataFrame([{"CV Type": cv_type, "Folds": cv_folds}])
        meta_df.to_excel(writer, sheet_name="Settings", index=False)
    return file_path


def run_cross_validation(
    df: pd.DataFrame,
    indicator_names,
    predictor_names,
    model,
    scaler: str,
    cv_type: str,
    cv_folds: int,
    useTransformer: str,
    transformer_cols,
    seed: int,
    problem_type: str,
    y_scaler_type: str = "none",
):
    if cv_type == "None" or cv_folds is None:
        return None, None

    if problem_type not in {"regression", "classification"}:
        return None, None

    X_raw = df[indicator_names].copy()
    y_raw = df[predictor_names].copy()

    if problem_type == "classification" and y_raw.shape[1] != 1:
        return None, None

    numeric_cols = X_raw.select_dtypes(include=np.number).columns.tolist()
    categorical_cols = []
    if useTransformer == "Yes" and transformer_cols is not None:
        categorical_cols = [c for c in transformer_cols.tolist() if c in X_raw.columns]

    preprocessor = make_preprocessor(numeric_cols=numeric_cols, categorical_cols=categorical_cols, cat_mode="onehot")
    steps = [("preprocess", preprocessor)]
    scaler_obj = _get_scaler(scaler)
    if scaler_obj is not None:
        steps.append(("scaler", scaler_obj))

    estimator = clone(model)
    if problem_type == "regression" and y_raw.shape[1] == 1 and y_scaler_type and y_scaler_type.lower() != "none":
        y_scaler = _get_scaler(y_scaler_type)
        if y_scaler is not None:
            estimator = TransformedTargetRegressor(regressor=estimator, transformer=y_scaler)

    steps.append(("model", estimator))
    pipeline = Pipeline(steps)

    folds = cv_folds if cv_folds and cv_folds >= 2 else 5
    if cv_type == "StratifiedKFold" and problem_type == "classification":
        cv = StratifiedKFold(n_splits=folds, shuffle=True, random_state=seed)
    elif cv_type == "RepeatedKFold":
        cv = RepeatedKFold(n_splits=folds, n_repeats=2, random_state=seed)
    elif cv_type == "RepeatedStratifiedKFold" and problem_type == "classification":
        cv = RepeatedStratifiedKFold(n_splits=folds, n_repeats=2, random_state=seed)
    elif cv_type == "ShuffleSplit":
        cv = ShuffleSplit(n_splits=folds, test_size=0.2, random_state=seed)
    elif cv_type == "StratifiedShuffleSplit" and problem_type == "classification":
        cv = StratifiedShuffleSplit(n_splits=folds, test_size=0.2, random_state=seed)
    else:
        cv = KFold(n_splits=folds, shuffle=True, random_state=seed)

    if problem_type == "classification":
        scoring = {
            "accuracy": "accuracy",
            "precision_weighted": "precision_weighted",
            "recall_weighted": "recall_weighted",
            "f1_weighted": "f1_weighted",
        }
    else:
        scoring = {
            "r2": "r2",
            "rmse": "neg_root_mean_squared_error",
            "mae": "neg_mean_absolute_error",
        }

    y_values = y_raw.iloc[:, 0] if isinstance(y_raw, pd.DataFrame) else y_raw
    results = cross_validate(pipeline, X_raw, y_values, scoring=scoring, cv=cv)
    summary_df, fold_df = _build_cv_summary(results)
    file_path = write_cross_validation_results(summary_df, fold_df, cv_type, folds)
    # Return both file path and summary data for frontend display
    return file_path, summary_df


def prediction(df_clean, best_model, training_features, X_scaler, y_scaler, feature_order, target_names=None):

    logger.info('Starting prediction function')

    # --- 1. Align features exactly as in training ---
    X_new = df_clean[feature_order]

    # --- 2. Apply X scaling IF it was used in training ---
    if X_scaler is not None:
        X_new_scaled = X_scaler.transform(X_new)
    else:
        X_new_scaled = X_new.values

    # --- 3. Predict ---
    y_pred = best_model.predict(X_new_scaled)

    # --- 4. Ensure 2D shape for inverse_transform ---
    y_pred = np.asarray(y_pred)
    original_shape = y_pred.shape
    if y_pred.ndim == 1:
        y_pred = y_pred.reshape(-1, 1)

    # --- 5. Inverse-transform ONLY if y_scaler exists ---
    if y_scaler is not None:
        y_pred = y_scaler.inverse_transform(y_pred)

    # --- 6. Handle single vs multiple targets ---
    num_targets = y_pred.shape[1] if y_pred.ndim > 1 else 1
    
    if num_targets > 1:
        # Multiple targets: create a column for each target
        if target_names is None:
            target_names = [f'Target_{i+1}' for i in range(num_targets)]
        elif len(target_names) != num_targets:
            # Fallback if target_names count doesn't match
            target_names = [f'Target_{i+1}' for i in range(num_targets)]
        
        for i, target_name in enumerate(target_names):
            pred_col = f'Predicted_{target_name}'
            df_clean[pred_col] = y_pred[:, i]
    else:
        # Single target: maintain backward compatibility
        y_pred_flat = y_pred.ravel()
        df_clean['Predicted_Target'] = y_pred_flat

    # --- 7. Save ---
    df_clean.to_csv(VIS_DIR / "predictions.csv", index=False)

    logger.info("Predictions complete. Saved to predictions.csv")

    
